"""文件解析业务逻辑服务"""

import logging
import json
import csv
import io
from typing import Dict, Any, List, Optional
from pathlib import Path
import struct

logger = logging.getLogger(__name__)


class FileParserService:
    """文件解析业务逻辑服务
    
    支持格式：
    - LAS (Log ASCII Standard)
    - CSV (Comma Separated Values)
    - XLS/XLSX (Excel spreadsheets)
    """

    @staticmethod
    def parse_las_file(file_content: bytes) -> Dict[str, Any]:
        """解析LAS格式文件
        
        LAS格式结构:
        - Header section: ~V2.0 或 ~V3.0
        - Well Information Section (~W)
        - Curve Information Section (~C)
        - Parameter Information Section (~P)
        - Other Information Section (~O)
        - ASCII Data Section (~A)
        """
        try:
            # 将二进制转换为文本
            text = file_content.decode('utf-8', errors='ignore')
            lines = text.split('\n')
            
            well_info = {}
            curve_info = []
            data_section_started = False
            data_lines = []
            
            current_section = None
            
            for i, line in enumerate(lines):
                line = line.strip()
                
                # 识别部分标记
                if line.startswith('~'):
                    current_section = line[1]
                    continue
                
                # 跳过空行和注释
                if not line or line.startswith('#'):
                    continue
                
                # 解析井信息部分
                if current_section == 'W':
                    if '.' in line:
                        parts = line.split(':')
                        if len(parts) >= 2:
                            key = parts[0].split('.')[0].strip()
                            value = ':'.join(parts[1:]).strip()
                            well_info[key] = value
                
                # 解析曲线信息部分
                elif current_section == 'C':
                    if '.' in line:
                        parts = line.split(':')
                        if len(parts) >= 2:
                            curve_name = parts[0].split('.')[0].strip()
                            curve_info.append({
                                "name": curve_name,
                                "mnemonic": curve_name,
                                "unit": parts[1].strip() if len(parts) > 1 else ""
                            })
                
                # 数据部分
                elif current_section == 'A':
                    if line and not line.startswith('~'):
                        data_lines.append(line)
            
            # 解析数据行
            data_values = []
            for line in data_lines:
                values = line.split()
                if len(values) == len(curve_info) + 1:  # +1 for depth
                    data_values.append([float(v) if v != '-999.25' else None for v in values])
            
            return {
                "success": True,
                "file_type": "LAS",
                "well_info": well_info,
                "curves": curve_info,
                "data_points": len(data_values),
                "data_sample": data_values[:10] if data_values else [],
                "message": f"LAS文件解析成功, 发现{len(curve_info)}条曲线, {len(data_values)}个数据点"
            }
        except Exception as e:
            logger.error(f"LAS文件解析失败: {str(e)}")
            return {
                "success": False,
                "error": "las_parse_failed",
                "message": f"LAS文件解析失败: {str(e)}"
            }

    @staticmethod
    def parse_csv_file(file_content: bytes) -> Dict[str, Any]:
        """解析CSV格式文件"""
        try:
            # 尝试多种编码
            encodings = ['utf-8', 'gbk', 'latin-1', 'utf-16']
            text = None
            
            for encoding in encodings:
                try:
                    text = file_content.decode(encoding)
                    break
                except:
                    continue
            
            if not text:
                return {
                    "success": False,
                    "error": "encoding_failed",
                    "message": "无法识别文件编码"
                }
            
            # 解析CSV
            csv_file = io.StringIO(text)
            reader = csv.reader(csv_file)
            
            headers = None
            data = []
            
            for i, row in enumerate(reader):
                if i == 0:
                    headers = [h.strip() for h in row]
                else:
                    data.append(row)
            
            if not headers:
                return {
                    "success": False,
                    "error": "no_headers",
                    "message": "CSV文件不包含表头"
                }
            
            return {
                "success": True,
                "file_type": "CSV",
                "headers": headers,
                "columns": len(headers),
                "data_points": len(data),
                "data_sample": data[:10] if data else [],
                "message": f"CSV文件解析成功, 发现{len(headers)}列, {len(data)}行数据"
            }
        except Exception as e:
            logger.error(f"CSV文件解析失败: {str(e)}")
            return {
                "success": False,
                "error": "csv_parse_failed",
                "message": f"CSV文件解析失败: {str(e)}"
            }

    @staticmethod
    def parse_excel_file(file_content: bytes, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """解析Excel格式文件"""
        try:
            import openpyxl
            from io import BytesIO
            
            # 识别Excel格式
            file_sig = file_content[:4]
            is_xlsx = file_sig == b'PK\x03\x04'  # XLSX是ZIP格式
            is_xls = file_sig == b'\xd0\xcf\x11\xe0'  # XLS是OLE2格式
            
            if not (is_xlsx or is_xls):
                return {
                    "success": False,
                    "error": "invalid_format",
                    "message": "无效的Excel文件格式"
                }
            
            # 使用openpyxl加载XLSX
            if is_xlsx:
                workbook = openpyxl.load_workbook(BytesIO(file_content))
                
                # 获取工作表
                if sheet_name:
                    if sheet_name not in workbook.sheetnames:
                        return {
                            "success": False,
                            "error": "sheet_not_found",
                            "message": f"工作表'{sheet_name}'不存在"
                        }
                    ws = workbook[sheet_name]
                else:
                    ws = workbook.active
                
                # 读取数据
                headers = []
                data = []
                
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:
                        headers = [str(h).strip() if h else f"Column{j}" for j, h in enumerate(row)]
                    else:
                        data.append(row)
                
                return {
                    "success": True,
                    "file_type": "XLSX",
                    "sheet_names": workbook.sheetnames,
                    "active_sheet": ws.title,
                    "headers": headers,
                    "columns": len(headers),
                    "data_points": len(data),
                    "data_sample": data[:10] if data else [],
                    "message": f"Excel文件解析成功, 发现{len(workbook.sheetnames)}个工作表, {len(headers)}列, {len(data)}行数据"
                }
            else:
                # XLS格式需要额外的库
                return {
                    "success": False,
                    "error": "xls_not_supported",
                    "message": "暂不支持XLS格式，请转换为XLSX"
                }
                
        except ImportError:
            return {
                "success": False,
                "error": "library_missing",
                "message": "需要openpyxl库支持Excel文件解析"
            }
        except Exception as e:
            logger.error(f"Excel文件解析失败: {str(e)}")
            return {
                "success": False,
                "error": "excel_parse_failed",
                "message": f"Excel文件解析失败: {str(e)}"
            }

    @staticmethod
    def detect_file_type(filename: str, file_content: bytes) -> Dict[str, Any]:
        """检测文件类型"""
        filename_lower = filename.lower()
        
        # 根据扩展名初步判断
        if filename_lower.endswith('.las'):
            return {"type": "LAS", "detected": True}
        elif filename_lower.endswith('.csv'):
            return {"type": "CSV", "detected": True}
        elif filename_lower.endswith(('.xlsx', '.xls')):
            return {"type": "EXCEL", "detected": True}
        
        # 根据文件签名判断
        if file_content.startswith(b'~V'):
            return {"type": "LAS", "detected": True}
        elif file_content.startswith(b'PK\x03\x04'):
            return {"type": "EXCEL", "detected": True, "format": "XLSX"}
        elif file_content.startswith(b'\xd0\xcf\x11\xe0'):
            return {"type": "EXCEL", "detected": True, "format": "XLS"}
        
        # 尝试作为文本格式
        try:
            text = file_content.decode('utf-8')
            if ',' in text:
                return {"type": "CSV", "detected": True}
        except:
            pass
        
        return {"type": "UNKNOWN", "detected": False}

    @staticmethod
    def parse_file(filename: str, file_content: bytes, **kwargs) -> Dict[str, Any]:
        """智能文件解析 - 自动检测格式并调用对应解析器"""
        
        # 限制文件大小 (100MB)
        MAX_SIZE = 100 * 1024 * 1024
        if len(file_content) > MAX_SIZE:
            return {
                "success": False,
                "error": "file_too_large",
                "message": f"文件过大，限制为{MAX_SIZE/1024/1024}MB"
            }
        
        # 检测文件类型
        type_info = FileParserService.detect_file_type(filename, file_content)
        
        if not type_info["detected"]:
            return {
                "success": False,
                "error": "unknown_format",
                "message": f"无法识别文件格式: {filename}"
            }
        
        file_type = type_info["type"]
        
        # 调用对应的解析器
        if file_type == "LAS":
            return FileParserService.parse_las_file(file_content)
        elif file_type == "CSV":
            return FileParserService.parse_csv_file(file_content)
        elif file_type == "EXCEL":
            sheet_name = kwargs.get("sheet_name")
            return FileParserService.parse_excel_file(file_content, sheet_name)
        else:
            return {
                "success": False,
                "error": "unsupported_format",
                "message": f"不支持的文件格式: {file_type}"
            }

    @staticmethod
    def validate_data_structure(parsed_data: Dict[str, Any]) -> Dict[str, Any]:
        """验证解析后的数据结构"""
        
        if not parsed_data.get("success"):
            return {
                "valid": False,
                "error": parsed_data.get("error", "parse_failed"),
                "message": parsed_data.get("message", "数据解析失败")
            }
        
        # 检查必要字段
        if "data_points" not in parsed_data:
            return {
                "valid": False,
                "error": "missing_data",
                "message": "数据不完整"
            }
        
        if parsed_data["data_points"] == 0:
            return {
                "valid": False,
                "error": "no_data",
                "message": "文件中没有数据"
            }
        
        # 检查列信息
        if parsed_data.get("file_type") == "LAS":
            if not parsed_data.get("curves"):
                return {
                    "valid": False,
                    "error": "no_curves",
                    "message": "LAS文件不包含曲线信息"
                }
        elif parsed_data.get("file_type") in ("CSV", "XLSX"):
            if not parsed_data.get("headers"):
                return {
                    "valid": False,
                    "error": "no_headers",
                    "message": "文件不包含列头信息"
                }
        
        return {
            "valid": True,
            "message": "数据结构验证通过"
        }
